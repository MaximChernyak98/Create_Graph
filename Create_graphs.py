from openpyxl import load_workbook
from openpyxl.chart import (LineChart, Reference, ScatterChart, Series, )


graphs_book = load_workbook('C:\\Work\\Python\\Create_graphs\\графики.xlsx')
ws = graphs_book.active


class MLine:
    num_x = 4
    num_y = 4
    line_name = ''

    def __init__(self, x_column, y_column, worksheet, style_tuple):
        self.x_column = x_column
        self.y_column = y_column
        self.style_tuple = style_tuple
        while worksheet.cell(row=self.num_x, column=self.x_column).value is not None:
            self.num_x += 1
        while worksheet.cell(row=self.num_y, column=self. y_column).value is not None:
            self.num_y += 1
        self.line_name = str(worksheet.cell(row=1, column=self.x_column).value)

    def create_plot(self):
        x_values = Reference(ws, min_col=self.x_column, min_row=4, max_row=self.num_x)
        y_values = Reference(ws, min_col=self.y_column, min_row=4, max_row=self.num_y)
        graph = Series(y_values, x_values, title=self.line_name)
        graph.marker.symbol = self.style_tuple[0]
        if self.style_tuple[1] is not None:
            graph.marker.graphicalProperties.solidFill = self.style_tuple[1]
        else:
            pass
        graph.marker.size = 7
        return graph

    def create_norms(self):
        x_values = Reference(ws, min_col=self.x_column, min_row=2, max_row=3)
        y_values = Reference(ws, min_col=self.y_column, min_row=2, max_row=3)
        norms = Series(y_values, x_values)
        norms.marker.symbol = 'triangle'
        norms.marker.size = 15
        norms.marker.graphicalProperties.solidFill = "FFFFFF"
        return norms


style_dict = {0: ('triangle', "FFFFFF"),
              1: ('triangle', None),
              2: ('square', "FFFFFF"),
              3: ('square', None),
              4: ('circle', "FFFFFF"),
              5: ('circle', None),
              6: ('diamond', "FFFFFF"),
              7: ('diamond', None),
              8: ('dot', "FFFFFF"),
              9: ('dot', None)
              }


def create_graphs(start_index, step, num_of_param, num_samples, chart, style_dict):
    current_offset = 0
    for sample in range(num_samples):
        line = MLine(x_column=(start_index+current_offset),
                     y_column=(start_index+current_offset+num_of_param),
                     worksheet=ws, style_tuple=style_dict.get(sample))
        chart.series.append(line.create_plot())
        current_offset += step



def number_to_letter(n):
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string


num_of_params = 6
step = 9


# Может нужен класс графиков?
for params in range(1, (num_of_params+1)):
    chart = ScatterChart()
    chart.title = None
    chart.style = 13
    chart.x_axis.title = 'Уровень воздействия, ед.'
    chart.y_axis.title = str(ws[f"{number_to_letter(params+1)}1"].value)
    create_graphs(start_index=1, step=step, num_of_param=params, num_samples=10, chart=chart, style_dict=style_dict)
    norms = MLine(x_column=1, y_column=1+params, worksheet=ws, style_tuple=('triangle', "FFFFFF"))
    chart.series.append(norms.create_norms())
    ws.add_chart(chart, f"{number_to_letter((params-1)*step+1)}21")



graphs_book.save('C:\\Work\\Python\\Create_graphs\\графики.xlsx')
graphs_book.close()