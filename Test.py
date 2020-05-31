import openpyxl
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import datetime, date, time
from openpyxl.chart import (LineChart, Reference, ScatterChart, Series, )
from openpyxl.chart.axis import DateAxis

graphs_book = load_workbook('C:\\Work\\Python\\Create_graphs\\графики.xlsx')
ws = graphs_book.active


print(ws["A1"].value)



graphs_book.save('C:\\Work\\Python\\Create_graphs\\графики.xlsx')
graphs_book.close()